#Original code from GitHub repo: https://github.com/lisaberrygis/AliasUpdater
#Edited and updated by Nicole Mucci, January 2025
#Included metadata pull from oracle and data storage in memory in excel file
#No other major changes to original code
#
# This script uses a lookup table to update alias names on a hosted feature service.
# The script updates the alias names in two places:
#   - The REST endpoint
#   - The layer's pop-up JSON via fieldInfos
#   - *If the layer was saved in the new Map Viewer in ArcGIS Online, updates the additional popupElement fieldInfos
# The pop-up configuration will not be altered with this implementation
# The script also allows you to update the long description, field type, and pop-up decimals/thousand separator for any field
#
# The script will use the input excel document and update any fields it finds that matches from the excel document.
# This script allows for multiple REST layers to be updated. Specify the REST layer count in the inputs.
# You must have ArcGIS Pro installed on your computer in order to run this script.
#
# Python version: 3.7 - Make sure your interpreter is calling to the arcgispro-py3 python.exe
# Updated: April 2020 - all http calls removed and replaced with python API calls
# Updated: July 2022 - Converted XLRD to OPENPYXL to read in excel file. XLRD no longer supports .xlsx files.
# Updated: August 2022 - can also update decimals for popup JSON.
#          Also updates popupElement in JSON if saved in new Map Viewer
# Updated: August 2023 - no longer need to input layer count, which is determined automatically. Also, blank values
#            in the excel doc are now handled by checking if they exist first, fixing NoneType error
# NOTE: As of 6/24, the script will alert you if you try to pass a long description with a < or > character. This will
#           not run as expected since the REST API cannot pass the characters to the service.

# Comments about inputs:_________________________________________________________________________________________
# username and password are your ArcGIS Online or ArcGIS Enterprise credentials
#
# layerID is the ID to a hosted feature service.
# *** You must own the service to run this script.
#
# lookupTable must be an excel document (.xlsx) with a header row.
#   The first column should be the field names
#   The second column should be the intended alias names for each field.
#   *optional* The third column should be the intended description for each field.
#   *optional* The fourth column can include the field type. This must be formatted
#           to match the backend JSON.
#           Ex:  nameOrTitle, description, typeOrCategory, countOrAmount, percentageOrRatio
#               measurement, currency, uniqueIdentifier, phoneNumber, emailAddress,
#               orderedOrRanked, binary, locationOrPlaceName, coordinate, dateAndTime
#   *optional* The fifth column can include a specification for how many decimals you want for each field
#               to have in the pop-up.
#   *optional* The sixth column can include a specification for if a numeric attribute should have a thousands comma
#                separator. Only specify this if it is a numeric field.
#               Ex: can use "true" or "yes" to specify. You can leave this column blank for any fields that are string
#                   or don't need a comma. You can also specify those as "no" or "false".
#
# If your script is having issues, make sure you at least have these 5 headers in the excel document,
# even if no values appear in the rows. This can cause the script to fail sometimes. Also make sure your excel file is closed.

# portalName can be left as-is if you are working in ArcGIS Online. Change to your portal URL otherwise.

# DATA PULL FROM ORACLE _______________________________________________________________________________________
survey_names = ["SCALLOP","HL","BTS","CSBLL","MMST","NARW","GOMBLL","SEAL","TURTLE","EDNA","OQ","SC", "SHRIMP", "COASTSPAN"]

import tempfile
import os
import pandas as pd
from io import BytesIO
import sqlalchemy
from sqlalchemy import create_engine, select, MetaData, Table
from sqlalchemy.orm import sessionmaker
from sqlalchemy.engine import create_engine
from arcgis.gis import GIS
from arcgis import gis
from arcgis.features import FeatureLayer
import openpyxl
from copy import deepcopy
import os
import copy

#Pull fields data from tables in oracle db
#CONNECT TO ORACLE
#Connect to oracle database using SQL alchemy engine
DIALECT = 'oracle'
SQL_DRIVER = 'cx_oracle'
USERNAME = '' #enter your username
PASSWORD = '' #enter your password
HOST = '' #enter the oracle db host url
PORT =  # enter the oracle port number
SERVICE = '' # enter the oracle db service name
ENGINE_PATH_WIN_AUTH = DIALECT + '+' + SQL_DRIVER + '://' + USERNAME + ':' + PASSWORD +'@' + HOST + ':' + str(PORT) + '/?service_name=' + SERVICE
engine = create_engine(ENGINE_PATH_WIN_AUTH)
connection= engine.connect()
Session = sessionmaker(bind=engine)
session= Session()

for survey_short in survey_names:
  
  #sql query to pull field names, aliases, and descriptions from oracle
  sql_query = f"SELECT DISTINCT COL_NAME, COL_ALIAS, COL_DESCRIPTION from SMIT_{survey_short}_META WHERE COL_NAME IS NOT NULL"
  df = pd.read_sql(sql_query, con=connection)
  #add two columns to dataframe, keep them null- field type and decimals
  #the code doesn't work without having these columns, see above description
  df['col_type'] = None
  df['col_dec'] = None 
  df['col_comma']= "no"
  #store the dataframe in memory as excel file
  excel_buffer = BytesIO()
  with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    
  excel_buffer.seek(0)
  
  #save in memory file to temp file
  with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
    temp_file.write(excel_buffer.read())
    temp_file_path = temp_file.name
  
  #extract the layer ID value from oracle metadata table using survey name
  layer_sql_query = f"SELECT FILE_ID from SMIT_{survey_short}_META"
  layerID = session.execute(layer_sql_query).fetchone()
  layerID = str(layerID[0])

# MAIN SCRIPT___________________________________________________________________________________________________
# Login to your arcgis account
  gis = GIS("PRO")
  portalName = "https://noaa.maps.arcgis.com/"
  # Get layer count from service
  updateItem = gis.content.get(layerID)
  restLayerCount = len(updateItem.layers)
  
  
  # format the path to the excel document so it is recognized as a path
  lookupTable = os.path.normpath(temp_file_path)
  
  # Read the lookup table and store the fields and alias names
  if lookupTable[-4:] != "xlsx":
      print("Please check your input. It needs to be a .xlsx excel file")
  else:
      print("Grabbing field and alias names from excel document...")
      # Open Master Metadata excel document
      workbook = openpyxl.load_workbook(lookupTable)
      sheet = workbook.active
      # Create an empty list to store all fields and alias names
      lookupList = []
  
      # Store values from master metadata excel doc and put into a list
      iter = sheet.iter_rows()
      iter.__next__()
      for row in iter:
          innerList = []
          for val in row:
              innerList.append(val.value)
          lookupList.append(innerList)

      looper = 0
      while restLayerCount > 0:
          # Access the feature layer intended for updating
          search = gis.content.search("id:" + layerID, item_type="Feature Layer")
          featureLayer = FeatureLayer.fromitem(search[0], layer_id=looper)
          layerName = search[0].name
          print("Updating layer " + str(looper) + " on " + str(layerName) + "...")
  
          print("\tGetting field definitions from service...")
          # Loop through fields in service and store JSON for any that are going to be updated
          layerFields = featureLayer.manager.properties.fields
  
          print("\tFinding fields to update...")
          # Loop through the fields in the service
          updateJSON = []
          for field in layerFields:
              fieldName = field['name']
              for lookupField in lookupList:
                  # As you loop through the service fields, see if they match a field in the excel document
                  if lookupField[0] == fieldName:
                      # store the field JSON from the online layer
                      fieldJSON = dict(deepcopy(field))
                      # assign the new alias name in JSON format
                      if lookupField[1]:
                          alias = lookupField[1]
                          fieldJSON['alias'] = alias
                      else:
                          alias = ""
                      # Assign field type, if specified
                      if lookupField[3]:
                          fldType = lookupField[3]
                      else:
                          fldType = ""
                      # assign the new field description in JSON format, if specified
                      if lookupField[2]:
                          longDesc = lookupField[2]
                          # Remove escape characters like double quotes, newlines, or encoding issues
                          if "<" or ">" in longDesc:
                              print("Special character > or < found in field: " + fieldName)
                              print("Script will not run as expected. Please remove all hyperlinks or > < characters from your long description and rerun the script.")
                          longDesc = longDesc.replace('"', '\\\"').replace("\n", " ").replace("\t", " ").replace(u'\xa0', u' ').replace(">=", " greater than or equal to ").replace("<=", " less than or equal to ").replace(">", " greater than ").replace("<", " less than ")
                      else:
                          longDesc = ""
                      # Build the JSON structure with the proper backslashes and quotes
                      fieldJSON['description'] = "{\"value\":" + "\"" + longDesc + "\"" + ",\"fieldValueType\":\"" + fldType + "\"}"
                      fieldJSON.pop('sqlType')
                      if alias != "":
                          print("\t\tField '" + fieldName + "' will be updated to get the alias name '" + alias + "'")
                      if longDesc != "":
                          print("\t\t\t\tThe long description for this field was also updated")
                      if fldType != "":
                          print("\t\t\t\tThe field type for this field was also updated to: " + fldType)
                      # Create a python list containing any fields to update
                      updateJSON.append(fieldJSON)
  
          if updateJSON:
              print("\tUpdating alias names of the REST service...")
              #jsonFormat =  json.dumps(updateJSON)
              aliasUpdateDict = {'fields': updateJSON}
              #aliasUpdateJSON = json.dumps(aliasUpdateDict)
              # Use the update definition call to push the new alias names into the service
              featureLayer.manager.update_definition(aliasUpdateDict)
              print("\tAlias names updated on service!")
  
          # Now check if the item has a pop-up configuration saving the alias names as well
          # First, grab the item JSON for the layer and create an item to hold the new edited JSON
          print("\tUpdating the alias names within the pop-up configuration on the item...")
          item = gis.content.get(layerID)
  
          # Grab the existing JSON for the popup, store a copy, and edit the aliases
          itemJSON = item.get_data(try_json=True)
          # Loop through the existing layer and check if any alias names don't match
          counter = 0
          if itemJSON:
              print("\tFinding all replacements of alias names within pop-up...")
              newItemJSON = copy.deepcopy(itemJSON)
              print("\t\tUpdating alias names in popup fieldInfos...")
              for i in itemJSON['layers'][looper]['popupInfo']['fieldInfos']:
                  fieldName2 = i['fieldName']
                  for lookup in lookupList:
                      if lookup[0] == fieldName2:
                          if lookup[1] != None:
                              newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['label'] = lookup[1]
                          # Check if there is a decimal spec
                          if "format" in i and "places" in i["format"]:
                              # If a value is specified in the lookup doc, assign that
                              if lookup[4] != None:
                                  newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['format']['places'] = lookup[4]
                              # If a value is not specified and the decimals have defaulted to 6, change to 2
                              else:
                                  if newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['format']['places'] == 6:
                                      newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['format']['places'] = 2
                          # Update thousands separator if lookup document specifies and if it exists in JSON
                          if lookup[5] != None and str(lookup[5]).lower() != "no" and str(lookup[5]).lower() != "false" and "format" in i and "digitSeparator" in i["format"]:
                              newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['format']['digitSeparator'] = True


                  counter += 1

              # Check if layer was updated in new Map Viewer and contains a popupElement JSON section with fieldInfos
              if "popupElements" in itemJSON['layers'][looper]['popupInfo'] and itemJSON['layers'][looper]['popupInfo']["popupElements"]:
                  c = 0
                  for i in itemJSON['layers'][looper]['popupInfo']["popupElements"]:
                      if i['type'] == 'fields':
                          print("\t\tUpdating popupElement fieldInfo...")
                          counter2 = 0
                          if "fieldInfos" in itemJSON['layers'][looper]['popupInfo']["popupElements"][c]:
                              for j in itemJSON['layers'][looper]['popupInfo']["popupElements"][c]["fieldInfos"]:
                                  fldName = j["fieldName"]
                                  for lkup in lookupList:
                                      if lkup[0] == fldName:
                                          if lkup[1] != None:
                                              newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['label'] = lkup[1]
                                          # Check if there is a decimal spec
                                          if "format" in j and "places" in j["format"]:
                                              # If a value is specified in the lookup doc, assign that
                                              if lkup[4] != None:
                                                  newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['format']['places'] = lkup[4]
                                              # If a value is not specified and the decimals have defaulted to 6, change to 2
                                              else:
                                                  if newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['format']['places'] == 6:
                                                      newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['format']['places'] = 2
                                          # Update thousands separator if lookup document specifies and if it exists in JSON
                                          if lkup[5] != None and str(lkup[5]).lower() != "no" and str(lkup[5]).lower() != "false" and "format" in j and "digitSeparator" in j["format"]:
                                              newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['format']['digitSeparator'] = True
                                  counter2 += 1
                      c += 1


              # Update json
              print("\tUpdating the alias names within the existing item pop-up...")
              portal = portalName
              update = item.update(item_properties={'text': newItemJSON})
              if update:
                  print("\tSuccess! Your alias names have been updated. Please check your service to confirm.")
              else:
                  print("\tUpdating pop-up failed.")
          else:
              print("\tNo pop-up JSON. Skipping.")

          looper += 1
          restLayerCount -= 1



print("Completed")        
os.remove(temp_file_path)

#some parts of the code wasn't working for ecomon survey, run this one separatly  
survey_names = ["ECOMON"] 

for survey_short in survey_names:
  
  #sql query to pull field names, aliases, and descriptions from oracle
  sql_query = f"SELECT COL_NAME, COL_ALIAS, COL_DESCRIPTION from SMIT_{survey_short}_META WHERE COL_NAME IS NOT NULL"
  df = pd.read_sql(sql_query, con=connection)
  #add two columns to dataframe, keep them null- field type and decimals
  #the code doesn't work without having these columns, see above description
  df['col_type'] = None
  df['col_dec'] = None 
  df['col_comma']= "no"
  #store the dataframe in memory as excel file
  excel_buffer = BytesIO()
  with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    
  excel_buffer.seek(0)
  
  #save in memory file to temp file
  with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
    temp_file.write(excel_buffer.read())
    temp_file_path = temp_file.name
  
  #extract the layer ID value from oracle metadata table using survey name
  layer_sql_query = f"SELECT FILE_ID from SMIT_{survey_short}_META"
  layerID = session.execute(layer_sql_query).fetchone()
  layerID = str(layerID[0])

# MAIN SCRIPT___________________________________________________________________________________________________
# Login to your arcgis account
  gis = GIS("PRO")
  portalName = "https://noaa.maps.arcgis.com/"
  # Get layer count from service
  updateItem = gis.content.get(layerID)
  restLayerCount = len(updateItem.layers)
  
  
  # format the path to the excel document so it is recognized as a path
  lookupTable = os.path.normpath(temp_file_path)
  
  # Read the lookup table and store the fields and alias names
  if lookupTable[-4:] != "xlsx":
      print("Please check your input. It needs to be a .xlsx excel file")
  else:
      print("Grabbing field and alias names from excel document...")
      # Open Master Metadata excel document
      workbook = openpyxl.load_workbook(lookupTable)
      sheet = workbook.active
      # Create an empty list to store all fields and alias names
      lookupList = []
  
      # Store values from master metadata excel doc and put into a list
      iter = sheet.iter_rows()
      iter.__next__()
      for row in iter:
          innerList = []
          for val in row:
              innerList.append(val.value)
          lookupList.append(innerList)

      looper = 0
      while restLayerCount > 0:
          # Access the feature layer intended for updating
          search = gis.content.search("id:" + layerID, item_type="Feature Layer")
          featureLayer = FeatureLayer.fromitem(search[0], layer_id=looper)
          layerName = search[0].name
          print("Updating layer " + str(looper) + " on " + str(layerName) + "...")
  
          print("\tGetting field definitions from service...")
          # Loop through fields in service and store JSON for any that are going to be updated
          layerFields = featureLayer.manager.properties.fields
  
          print("\tFinding fields to update...")
          # Loop through the fields in the service
          updateJSON = []
          for field in layerFields:
              fieldName = field['name']
              for lookupField in lookupList:
                  # As you loop through the service fields, see if they match a field in the excel document
                  if lookupField[0] == fieldName:
                      # store the field JSON from the online layer
                      fieldJSON = dict(deepcopy(field))
                      # assign the new alias name in JSON format
                      if lookupField[1]:
                          alias = lookupField[1]
                          fieldJSON['alias'] = alias
                      else:
                          alias = ""
                      # Assign field type, if specified
                      if lookupField[3]:
                          fldType = lookupField[3]
                      else:
                          fldType = ""
                      # assign the new field description in JSON format, if specified
                      if lookupField[2]:
                          longDesc = lookupField[2]
                          # Remove escape characters like double quotes, newlines, or encoding issues
                          if "<" or ">" in longDesc:
                              print("Special character > or < found in field: " + fieldName)
                              print("Script will not run as expected. Please remove all hyperlinks or > < characters from your long description and rerun the script.")
                          longDesc = longDesc.replace('"', '\\\"').replace("\n", " ").replace("\t", " ").replace(u'\xa0', u' ').replace(">=", " greater than or equal to ").replace("<=", " less than or equal to ").replace(">", " greater than ").replace("<", " less than ")
                      else:
                          longDesc = ""
                      # Build the JSON structure with the proper backslashes and quotes
                      fieldJSON['description'] = "{\"value\":" + "\"" + longDesc + "\"" + ",\"fieldValueType\":\"" + fldType + "\"}"
                      fieldJSON.pop('sqlType')
                      if alias != "":
                          print("\t\tField '" + fieldName + "' will be updated to get the alias name '" + alias + "'")
                      if longDesc != "":
                          print("\t\t\t\tThe long description for this field was also updated")
                      if fldType != "":
                          print("\t\t\t\tThe field type for this field was also updated to: " + fldType)
                      # Create a python list containing any fields to update
                      updateJSON.append(fieldJSON)
  
          if updateJSON:
              print("\tUpdating alias names of the REST service...")
              #jsonFormat =  json.dumps(updateJSON)
              aliasUpdateDict = {'fields': updateJSON}
              #aliasUpdateJSON = json.dumps(aliasUpdateDict)
              # Use the update definition call to push the new alias names into the service
              featureLayer.manager.update_definition(aliasUpdateDict)
              print("\tAlias names updated on service!")
  
          # Now check if the item has a pop-up configuration saving the alias names as well
          # First, grab the item JSON for the layer and create an item to hold the new edited JSON
          print("\tUpdating the alias names within the pop-up configuration on the item...")
          item = gis.content.get(layerID)
  
          # Grab the existing JSON for the popup, store a copy, and edit the aliases
          itemJSON = item.get_data(try_json=True)
          # Loop through the existing layer and check if any alias names don't match
          counter = 0
          if itemJSON:
              print("\tFinding all replacements of alias names within pop-up...")
              newItemJSON = copy.deepcopy(itemJSON)
              print("\t\tUpdating alias names in popup fieldInfos...")
              # for i in itemJSON['layers'][looper]['popupInfo']['fieldInfos']:
              #     fieldName2 = i['fieldName']
              #     for lookup in lookupList:
              #         if lookup[0] == fieldName2:
              #             if lookup[1] != None:
              #                 newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['label'] = lookup[1]
              #             # Check if there is a decimal spec
              #             if "format" in i and "places" in i["format"]:
              #                 # If a value is specified in the lookup doc, assign that
              #                 if lookup[4] != None:
              #                     newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['format']['places'] = lookup[4]
              #                 # If a value is not specified and the decimals have defaulted to 6, change to 2
              #                 else:
              #                     if newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['format']['places'] == 6:
              #                         newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['format']['places'] = 2
              #             # Update thousands separator if lookup document specifies and if it exists in JSON
              #             if lookup[5] != None and str(lookup[5]).lower() != "no" and str(lookup[5]).lower() != "false" and "format" in i and "digitSeparator" in i["format"]:
              #                 newItemJSON['layers'][looper]['popupInfo']['fieldInfos'][counter]['format']['digitSeparator'] = True
              # 
              # 
              #     counter += 1
              # 
              # # Check if layer was updated in new Map Viewer and contains a popupElement JSON section with fieldInfos
              # if "popupElements" in itemJSON['layers'][looper]['popupInfo'] and itemJSON['layers'][looper]['popupInfo']["popupElements"]:
              #     c = 0
              #     for i in itemJSON['layers'][looper]['popupInfo']["popupElements"]:
              #         if i['type'] == 'fields':
              #             print("\t\tUpdating popupElement fieldInfo...")
              #             counter2 = 0
              #             if "fieldInfos" in itemJSON['layers'][looper]['popupInfo']["popupElements"][c]:
              #                 for j in itemJSON['layers'][looper]['popupInfo']["popupElements"][c]["fieldInfos"]:
              #                     fldName = j["fieldName"]
              #                     for lkup in lookupList:
              #                         if lkup[0] == fldName:
              #                             if lkup[1] != None:
              #                                 newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['label'] = lkup[1]
              #                             # Check if there is a decimal spec
              #                             if "format" in j and "places" in j["format"]:
              #                                 # If a value is specified in the lookup doc, assign that
              #                                 if lkup[4] != None:
              #                                     newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['format']['places'] = lkup[4]
              #                                 # If a value is not specified and the decimals have defaulted to 6, change to 2
              #                                 else:
              #                                     if newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['format']['places'] == 6:
              #                                         newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['format']['places'] = 2
              #                             # Update thousands separator if lookup document specifies and if it exists in JSON
              #                             if lkup[5] != None and str(lkup[5]).lower() != "no" and str(lkup[5]).lower() != "false" and "format" in j and "digitSeparator" in j["format"]:
              #                                 newItemJSON['layers'][looper]['popupInfo']['popupElements'][c]["fieldInfos"][counter2]['format']['digitSeparator'] = True
              #                     counter2 += 1
              #         c += 1


              # Update json
              print("\tUpdating the alias names within the existing item pop-up...")
              portal = portalName
              update = item.update(item_properties={'text': newItemJSON})
              if update:
                  print("\tSuccess! Your alias names have been updated. Please check your service to confirm.")
              else:
                  print("\tUpdating pop-up failed.")
          else:
              print("\tNo pop-up JSON. Skipping.")

          looper += 1
          restLayerCount -= 1


print("Completed")        
os.remove(temp_file_path)
